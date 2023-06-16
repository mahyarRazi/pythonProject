from openpyxl import load_workbook
from hashlib import sha512
import xlrd
import xlwt
import requests
hbj
file = load_workbook('entryData.xlsx')
activeSheet = file['Data']
for i in range(activeSheet.max_row):
  import requests

  url = "https://devbpms.okcs.com/WebServices/WorkflowEngineSOA"

  payload = "<?xml version=\"1.0\" encoding=\"utf-8\"?>\r\n<soap:Envelope xmlns:xsi=\"http://www.w3.org/2001/XMLSchema-instance\" xmlns:xsd=\"http://www.w3.org/2001/XMLSchema\" xmlns:soap=\"http://schemas.xmlsoap.org/soap/envelope/\">\r\n  <soap:Body>\r\n    <createCasesAsString xmlns=\"http://tempuri.org/\">\r\n      <casesInfo>\r\n         <![CDATA[<BizAgiWSParam>\r\n\r\n         <domain>okco.ir</domain>\r\n\r\n         <userName>"+str(activeSheet.cell(row=i+1, column=1).value)+"</userName>\r\n\r\n         <Cases>\r\n\r\n                 <Case>\r\n\r\n                 <Process>soratJalaseh</Process>\r\n\r\n                 <Entities>\r\n\r\n                         <SoratjalaseRequest>\r\n                                 <epmSn>"+str(activeSheet.cell(row=i+1, column=2).value)+"</epmSn>\r\n                                 <JalaseTitle>"+str(activeSheet.cell(row=i+1, column=3).value)+"</JalaseTitle>\r\n                                 <JalaseSubject>"+str(activeSheet.cell(row=i+1, column=4).value)+"</JalaseSubject>\r\n                                 <SoratjalaseUserConfirm>"+str(activeSheet.cell(row=i+1, column=5).value)+"</SoratjalaseUserConfirm>\r\n                                 <HuzeJalase>"+str(activeSheet.cell(row=i+1, column=6).value)+"</HuzeJalase>\r\n                                <JalaseDate>"+str(activeSheet.cell(row=i+1, column=7).value)+"</JalaseDate>\r\n                                <vaziyatSoratJalase>"+str(activeSheet.cell(row=i+1, column=8).value)+"</vaziyatSoratJalase>\r\n                         </SoratjalaseRequest>\r\n\r\n                 </Entities>\r\n\r\n                 </Case>\r\n\r\n         </Cases>\r\n\r\n </BizAgiWSParam>]]>\r\n      </casesInfo>\r\n    </createCasesAsString>\r\n  </soap:Body>\r\n</soap:Envelope>"
  payload = payload.encode('utf-8')
  headers = {
    'Host': 'devbpms.okcs.com',
    'Content-Type': 'text/xml; charset=utf-8',
    'SOAPAction': '"http://tempuri.org/createCasesAsString"'
  }

  response = requests.request("POST", url, headers=headers, data=payload)

  print(response.text)

