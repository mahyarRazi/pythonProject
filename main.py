from openpyxl import load_workbook
from hashlib import sha512
import xlrd
import xlwt
import requests

file = load_workbook('entryData.xlsx')
activeSheet = file['Data']
#print((activeSheet.cell( row=1, column=1)).value)
# print(activeSheet.max_row)



# passValue = (activeSheet.cell(row=1, column=1)).value
# passwoard = ("CRYPT.2.BIZAGI:"+(sha512(str(passValue).encode()).hexdigest()).upper())


for i in range(activeSheet.max_row):
  import requests

  url = "https://devbpms.okcs.com/WebServices/entitymanagersoa.asmx"

  payload = "<?xml version=\"1.0\" encoding=\"utf-8\"?>\r\n<soap:Envelope xmlns:xsi=\"http://www.w3.org/2001/XMLSchema-instance\" xmlns:xsd=\"http://www.w3.org/2001/XMLSchema\" xmlns:soap=\"http://schemas.xmlsoap.org/soap/envelope/\">\r\n    <soap:Body>\r\n        <saveEntityAsString xmlns=\"http://tempuri.org/\">\r\n            <entityInfo>\r\n\r\n\r\n\r\n  <![CDATA[<BizAgiWSParam>\r\n\r\n     <Entities>\r\n\r\n\t\t <KarshenashaTarifKalaOld>\r\n\r\n\t\t\t<RecId>"+str(activeSheet.cell(row=i+1, column=1).value)+"</RecId>\r\n\r\n\t\t</KarshenashaTarifKalaOld>\r\n\r\n     </Entities>\r\n\r\n  </BizAgiWSParam>]]>\r\n\r\n\r\n            </entityInfo>\r\n        </saveEntityAsString>\r\n    </soap:Body>\r\n</soap:Envelope>"
  headers = {
    'Host': 'devbpms.okcs.com',
    'Content-Type': 'text/xml; charset=utf-8',
    'SOAPAction': '"http://tempuri.org/saveEntityAsString"'
  }

  response = requests.request("POST", url, headers=headers, data=payload)

  print(response.text)




  # import requests

  # url = "https://devbpms.okcs.com/WebServices/entitymanagersoa.asmx"
  #
  # payload = "<?xml version=\"1.0\" encoding=\"utf-8\"?>\r\n<soap:Envelope xmlns:xsi=\"http://www.w3.org/2001/XMLSchema-instance\" xmlns:xsd=\"http://www.w3.org/2001/XMLSchema\" xmlns:soap=\"http://schemas.xmlsoap.org/soap/envelope/\">\r\n    <soap:Body>\r\n        <saveEntityAsString xmlns=\"http://tempuri.org/\">\r\n            <entityInfo>\r\n                <![CDATA[<BizAgiWSParam><Entities><ListKalaNoQC><RecID>"+str(activeSheet.cell(row=i+1, column=1).value)+"\r\n</RecID></ListKalaNoQC>\r\n\r\n     </Entities>\r\n\r\n  </BizAgiWSParam>]]>\r\n\r\n\r\n        \r\n    \r\n    </entityInfo>\r\n</saveEntityAsString>\r\n</soap:Body>\r\n</soap:Envelope>"
  # headers = {
  #     'Host': 'devbpms.okcs.com',
  #     'Content-Type': 'text/xml; charset=utf-8',
  #     'SOAPAction': '"http://tempuri.org/saveEntityAsString"'
  # }
  #
  # response = requests.request("POST", url, headers=headers, data=payload)
  #
  # print(response.text)




