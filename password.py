from openpyxl import load_workbook
from hashlib import sha512
from openpyxl import workbook
import xlrd
import xlwt
import requests

file = load_workbook('entryData.xlsx')
activeSheet = file['Data']
#print((activeSheet.cell( row=1, column=1)).value)
print(activeSheet.max_row)


# passValue = (activeSheet.cell(row=1, column=1)).value
# passwoard = ("CRYPT.2.BIZAGI:"+(sha512(str(passValue).encode()).hexdigest()).upper())


for i in range(activeSheet.max_row):
    passValue = (activeSheet.cell(row=i+1, column=1)).value
    passwoard = ("CRYPT.2.BIZAGI:" + (sha512(str(passValue).encode()).hexdigest()).upper())

    activeSheet.cell(row=i+1, column=2).value = passwoard
file.save("DataFOrImport.xlsx")




